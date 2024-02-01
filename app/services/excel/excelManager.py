from app.services.excel.excelToJSON import ExceltoJson 
from app.services.excel.updateExcel import UpdateExcel
from app.services.excel.columnsToDB import columnsTODB
from app.shared.constantsAPI import Config

from openpyxl.utils import get_column_letter
import openpyxl


'''
La siguiente clase tendra la funcionalidad de recorrer el excel en el cual se
encuentran los datos de los inquilinos.

Ojo : Es necesario que entienda la estructura del excel para entender las
siguientes funciones.
'''

class ExcelManager:
    def __init__(self, file):
        # Lectura del excel 
        self.excel = openpyxl.load_workbook(file)
        self.hojas = self.excel.get_sheet_names()
        
        # Clases para manejo de informacion
        self.excelToJSON = ExceltoJson()
        self.UpdateExcel = UpdateExcel()
        self.trazables = {}

        # Atributos del excel
        self.columnaInicial = 3
        self.filaInicial = 5
        self.filaInicialGanado = 6
        
    def ExcelTrazablesUpdateToDB(self):
        
        '''
        Se van a recorrer las hojas del excel, y se va a crear la lista
        de elementos que van a ser subidos en los end-Points 

        Es importante el orden de lectura de las hojas debido a las ubicaciones
        y lo padres.
        '''

        self.get_dataHojaTrazable("FINCAS")
        self.get_dataHojaTrazable("POTREROS")
        self.get_dataHojaTrazable("LOTES")
        self.get_dataHojaTrazable("GANADO")

        ubicaciones = self.excelToJSON.JsonUbicacion(self.trazables)
        
        #print("------------------")
        #print("Diccionario Ubicaciones")
        #print("------------------\n")
        #print(self.trazables["FIN"])
        #print("------------------\n")
        
        self.UpdateExcel.ubicaciones = ubicaciones
        self.UpdateExcel.updateInformation()
        
        #print("------------------")
        #print("Inventarios")
        #print("------------------\n")

        #for inv in self.UpdateExcel.inventarios:
        #    print(inv)

        #print("------------------")
        #print("Datos")
        #print("------------------\n")
        #for dat in self.UpdateExcel.datos:
        #    print(dat)


    def get_dataHojaTrazable(self, nombreHoja):
        
        '''
        Se va a recorrer la hoja de exel extrayendo la cantidad de filas
        atravez de la libreria openpyxl, mietras que la cantidad de columnas la 
        extraemos dependiendo de hoja que se esta recorriendo. 

        La fila extraida se va a pasar diferentes funciones con para que sea 
        procesada y transformada en los Json para ser subidos a la API. 
        '''

        information = self.excel.get_sheet_by_name(nombreHoja)

        cantOfRows = len(list(information.rows))
        getCantOfColumns = self.getCantOfColumns(nombreHoja)
        maxColumns = self.columnaInicial + getCantOfColumns 
        
        headers, columnasObligatorias = self.getHeaders(information,maxColumns,nombreHoja)
        
        filaInicial = self.filaInicial
        if nombreHoja == "GANADO":
            filaInicial = self.filaInicialGanado

        for row in range(filaInicial, cantOfRows):
            rowData = [] 
            i = 0
            
            for column in range(self.columnaInicial, maxColumns):
                columnChar = get_column_letter(column)
                value = information[columnChar + str(row)].value
                rowData.append(value)
            
            if self.IsemptyRow(rowData):
                break

            dictInvetario = self.excelToJSON.JsonInventario(
                rowData, nombreHoja, self.trazables
                )

            uuid = dictInvetario['id_trazii']
            
            dictDatos = self.excelToJSON.JsonDatos(
                rowData, headers, nombreHoja, uuid
            )

            padre = self.setTrazable(dictInvetario, headers, rowData)

            if padre != None:
                dictInvetario['padre'] = padre

            self.UpdateExcel.inventarios.append(dictInvetario)
            self.UpdateExcel.datos.extend(dictDatos)
        
    def print_data(self, nombreHoja):
        
        '''
        Se va a recorrer la hoja de exel extrayendo la cantidad de filas 
        atravez de la libreria, mietras que la cantidad de columnas la 
        extraemos dependiendo de hoja que se esta recorriendo. 
        '''

        information = self.excel.get_sheet_by_name(nombreHoja)
        cantOfRows = len(list(information.rows))
        getCantOfColumns = self.getCantOfColumns(nombreHoja)
        maxColumns = self.columnaInicial + getCantOfColumns 
        
        print("Cantidad de filas : ")
        print(cantOfRows)
        print("Cantidad de columnas : ")
        print(maxColumns)
       
        filaInicial = self.filaInicial
        if nombreHoja == "GANADO":
            filaInicial = self.filaInicialGanado

        Datos = False
        for row in range(filaInicial - 1 ,cantOfRows):
            for column in range(self.columnaInicial, maxColumns):
                columnChar = get_column_letter(column)
                value = information[columnChar + str(row)].value
                print("cell "+columnChar + str(row), end= " : ")
                print(str(value) + " | ", end= " ")
            print()

    def setTrazable(self, dictInvetario, headers, rowData):
        
        '''
        Los datos del excel que se van recorriendo van a ser guardados en un
        diccionario (como atributo de la clase) con el fin de ubicar su padre
        y su ubicacion.
        
        Actualmente la unica ubicacicon es finca.
        '''
        
        trazable = dictInvetario['trazable']
        denominacion = dictInvetario['denominacion']
        id_trazii = dictInvetario['id_trazii']
        ubicacion = self.get_Ubicacion(headers, rowData, trazable, id_trazii)
        padre = self.get_Padre(headers, rowData, trazable, id_trazii)
        ubicacion_denominacion = None

        if ubicacion != None:
            if ubicacion[1]:
                ubicacion_denominacion = ubicacion[0]['denominacion']
                ubicacion = ubicacion[0]['id_trazii']
            else:
                ubicacion_denominacion = ubicacion[0]['ubicacion_denominacion']
                ubicacion = ubicacion[0]['ubicacion']

        if padre != None:  
            padre = padre['id_trazii']

        if trazable not in self.trazables:
            self.trazables[trazable] = {}
        
        self.trazables[trazable][denominacion] = {
            'id_trazii' : id_trazii,
            'denominacion' : denominacion,
            'ubicacion' : ubicacion,
            'ubicacion_denominacion' : ubicacion_denominacion,
            'padre' : padre,
            'elementos' : [],
            'hijos' : []
        }

        return padre
    
    def get_Padre(self, headers, rowData, trazable, id_trazii):
        
        '''
        Esta funcion inicia verificando el tipo en de padre del trazable. 

        A continuacion, se realiza la busqueda de la denominacion asociada
        a al padre del elemento.

        Finalmente el tipo de trazable padre y su denominacion se encuentra
        su uuid dentro del diccionario de trazbles.   
        '''
                
        padres = columnsTODB.padres
        padre = None

        for key in padres:
            if trazable == key:
                padre = padres[key]
                break
        
        if padre == None:
            return None
        
        cantOfCells = len(rowData)
        denominacion = ""
        
        if padre == Config.Trazables.FINCA:   
            Column = "*Finca"
        elif padre == Config.Trazables.GANADO:
            Column = "*Lote"
                        
        # Si tiene el nombre de la finca, el nombre de esta.
        i = 0
        while i < cantOfCells:
            if headers[i] == Column:
                denominacion = rowData[i]
                break
            i += 1
        
        #print("padre : " + str(padre)) 
        #print("denominacion : "+str(denominacion))
        self.trazables[padre][denominacion]['hijos'].append(id_trazii)
        return self.trazables[padre][denominacion]
    
    def get_Ubicacion(self, headers, rowData, trazable, id_trazii):
        
        '''
        Esta funcion inicia verificando el tipo en de ubicacion del trazable. 

        A continuacion, se realiza la busqueda de la denominacion asociada
        a ala ubicacion del elemento.

        Finalmente el tipo de trazable ubicacion y su denominacion se encuentra
        su uuid dentro del diccionario de trazbles.   

        Esta funcion retorna, el trazable que contiene la ubicacion del elemento
        y un boleano que indica si, su ubicacion era implicita en el excel.
        '''

        ubicaciones = columnsTODB.ubicaciones
        ubicacion = None
        
        for key in ubicaciones:
            if trazable in ubicaciones[key]:
                ubicacion = key
                break
        
        if ubicacion == None:
            return None
        
        cantOfCells = len(rowData)
        denominacion = ""

        if ubicacion == Config.Trazables.FINCA:   
            Column = "*Finca"

        # Verificar si la hoja de excel tiene el nombre de la finca.
        haveFinca = (Column in headers)
        
        # Si no tiene la finca, se verifica donde esta ubicado su padre.
        if not haveFinca:
            padre = self.get_Padre(headers, rowData, trazable, id_trazii)
            denominacion = padre['ubicacion_denominacion']
            self.trazables[ubicacion][denominacion]['elementos'].append(id_trazii)
            return [padre, False]
        
        # Si tiene el nombre de la finca, el nombre de esta.
        i = 0
        while i < cantOfCells and haveFinca:
            if headers[i] == Column:
                denominacion = rowData[i]
                break
            i += 1

        #print("ubicacion : " + str(ubicacion)) 
        #print("denominacion : "+str(denominacion))

        self.trazables[ubicacion][denominacion]['elementos'].append(id_trazii)
        return [self.trazables[ubicacion][denominacion], True]
    
    def getHeaders(self, information, maxColumns, nombreHoja):
                
        '''
        Lee la hoja de excel seleccionada, ingresa el nombre de las columnas
        en un arreglo y va indicando en un arreglo de boleanos si ese dato es
        obligatorio.
        '''

        headers = []
        columnasObligatorias = []
        
        filaInicial = self.filaInicial
        if nombreHoja == "GANADO":
            filaInicial = self.filaInicialGanado

        for column in range(self.columnaInicial, maxColumns):
            columnChar = get_column_letter(column)
            value = information[columnChar + str(filaInicial - 1)].value
            value = str(value.strip())

            if value.startswith("*"):
                columnasObligatorias.append(True)
            else:
                columnasObligatorias.append(False)
            
            headers.append(value)
        
        return (headers, columnasObligatorias)
    
    def getCantOfColumns(self, nombreHoja):
        if nombreHoja == "FINCAS":
            return 5
        if nombreHoja == "POTREROS":
            return 6
        if nombreHoja == "LOTES":
            return 3
        if nombreHoja == "GANADO":
            return 21
        
    def IsemptyRow(self, rowData):
        cantOfCells = len(rowData)
        cantEmptyCell = 0

        for cell in rowData:
            if cell == None:
                cantEmptyCell += 1

        if cantEmptyCell == cantOfCells:
            return True
        
        return False
